"""
FL Studio Project Name Manager
Excel exporter — writes scanned project data to an .xlsx file.
"""

from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fl_project_scanner import FLProject


# ── Style constants ──────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2B2B3D", end_color="2B2B3D", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

NAMED_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
UNNAMED_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
EMPTY_FILL = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")

ROW_FONT = Font(name="Segoe UI", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D5D8DC"),
    right=Side(style="thin", color="D5D8DC"),
    top=Side(style="thin", color="D5D8DC"),
    bottom=Side(style="thin", color="D5D8DC"),
)


def export_to_excel(projects: List[FLProject], output_path: str) -> str:
    """
    Export the list of scanned projects to an Excel workbook.

    Parameters
    ----------
    projects : list[FLProject]
        The projects to export.
    output_path : str
        Full path (including filename) for the .xlsx file.

    Returns
    -------
    str
        The absolute path of the saved file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "FL Studio Projects"

    # ── Headers ──────────────────────────────────────────────────────────
    headers = [
        "#",
        "Folder Name",
        "Project Name(s)",
        "FLP Count",
        "Audio Files",
        "Status",
        "Full Path",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # ── Data rows ────────────────────────────────────────────────────────
    for row_idx, project in enumerate(projects, start=2):
        num = row_idx - 1
        flp_names = ", ".join(Path(f).stem for f in project.flp_files)
        audio_names = ", ".join(project.audio_files) if project.audio_files else "—"

        values = [
            num,
            project.folder_name,
            flp_names,
            project.project_count,
            audio_names,
            project.status,
            project.folder_path,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = ROW_FONT
            cell.border = THIN_BORDER

            # Color-code the status column
            if col_idx == 6:
                if project.status == "Named ✓":
                    cell.fill = NAMED_FILL
                elif project.status == "Unnamed":
                    cell.fill = UNNAMED_FILL
                else:
                    cell.fill = EMPTY_FILL

    # ── Auto-fit column widths (approximate) ─────────────────────────────
    col_widths = [6, 30, 45, 12, 35, 14, 60]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze the header row
    ws.freeze_panes = "A2"

    # ── Summary sheet ────────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    total = len(projects)
    named = sum(1 for p in projects if p.status == "Named ✓")
    unnamed = sum(1 for p in projects if p.status == "Unnamed")
    empty = sum(1 for p in projects if p.status == "Empty")

    summary_data = [
        ("Total Projects", total),
        ("Named (has audio)", named),
        ("Unnamed (no audio)", unnamed),
        ("Empty Folders", empty),
    ]
    for r, (label, val) in enumerate(summary_data, start=1):
        c1 = ws_summary.cell(row=r, column=1, value=label)
        c1.font = Font(name="Segoe UI", size=11, bold=True)
        c2 = ws_summary.cell(row=r, column=2, value=val)
        c2.font = Font(name="Segoe UI", size=11)

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 12

    # ── Save ─────────────────────────────────────────────────────────────
    wb.save(output_path)
    return str(Path(output_path).resolve())
